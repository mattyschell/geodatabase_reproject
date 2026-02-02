import arcpy
import sys
import logging
import os
import openpyxl
import time
import shutil
import stat

from filegeodatabasemanager import localgdb 


# arcpy.topographic
# requires an Esri ArcGIS Production Mapping license

# This code was originally developed by Chris S. and colleagues @ ESRI
# He copied and donatepasted the orinal "reproject" submodule
# I think ESRI developed that sub to ingest CSCL into their locators?
# We have modified and wrapped up reproject here
# We will use this to correct tolerance and resolution issues in CSCL
# migrating from legacy CSCL data (with class extensions) to modern CSCL
       

def reproject (
    gdbin
   ,gdbout
   ,logger
   ,workdir
   ,srid
   ,input_spec_xlsx: str | None = None
   ,create_xlsx: bool = True
) -> None:

    """Reproject CSCL file geodatabase

    Args:
        cscl_in_fgdb (str): Path to input file geodatabase
        cscl_out_fgdb (str): Path to output file geodatabase
        input_spec_xlsx (str | None, optional): Path to intermediate Excel file (written by this function)
        object_map_fgdb (str | None, optional): Path to intermediate file geodatabase (written by this function)
        create_xlsx (bool, optional): Whether to create the intermediate Excel file from scratch. Defaults to True.
    """

    logger.info('calling reproject on {0} to {1} with srid {2}'.format(gdbin.name
                                                                      ,gdbout.name
                                                                      ,srid))

    if not input_spec_xlsx:
        input_spec_xlsx = os.path.join(workdir
                                      ,"CorrectedProjection.xlsx")  # type: ignore

    object_map_gdb = localgdb(os.path.join(workdir
                                          ,"CSCL_OLD_TO_NEW.gdb"))  # type: ignore

    logger.info(
        f"Processing started to Re-project {gdbin.name} to {gdbout.name}"
    )
    logger.debug(f"{input_spec_xlsx=}")

     
    if create_xlsx:

        # Generate Excel from Geodatabase
        logger.info(f"Generating Excel from Geodatabase {gdbin.gdb}")
        temp_xlsx = os.path.join(workdir
                                ,"OriginalProjection.xlsx")  # type: ignore
        
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

        arcpy.topographic.GenerateExcelFromGeodatabase(gdbin.gdb, temp_xlsx)

        # Load Excel into openpyxl and update spatial reference
        logger.info("Updating Spatial Reference in Excel file")
        wb = openpyxl.load_workbook(temp_xlsx)

        # Overwrite SpatialReferences sheet
        ws_sr = wb["SpatialReferences"]
        for row in ws_sr.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        ws_sr["A2"] = 1

        if int(srid) == 2263:
            
            ws_sr["B2"] = "NAD_1983_StatePlane_New_York_Long_Island_FIPS_3104_Feet"
            ws_sr["C2"] = 2263
            ws_sr["D2"] = (  # ? Missing last line of constants.NY_STATE_PLANE?
                'PROJCS["NAD_1983_StatePlane_New_York_Long_Island_FIPS_3104_Feet",GEOGCS["GCS_North_American_1983",'
                'DATUM["D_North_American_1983",SPHEROID["GRS_1980",6378137.0,298.257222101]],PRIMEM["Greenwich",0.0],'
                'UNIT["Degree",0.0174532925199433]],PROJECTION["Lambert_Conformal_Conic"],'
                'PARAMETER["False_Easting",984250.0],PARAMETER["False_Northing",0.0],PARAMETER["Central_Meridian",-74.0],'
                'PARAMETER["Standard_Parallel_1",40.66666666666666],PARAMETER["Standard_Parallel_2",41.03333333333333],'
                'PARAMETER["Latitude_Of_Origin",40.16666666666666],UNIT["Foot_US",0.3048006096012192]]'
            )

        elif int(srid) == 6539:

            # mschell added this do not blame ESRI if it doesnt work

            ws_sr["B2"] = "NAD_1983_2011_StatePlane_New_York_Long_Isl_FIPS_3104_Ft_US"
            ws_sr["C2"] = 6539
            ws_sr["D2"] = ( 
                'PROJCS["NAD_1983_2011_StatePlane_New_York_Long_Isl_FIPS_3104_Ft_US",GEOGCS["GCS_NAD_1983_2011",'
                'DATUM["D_NAD_1983_2011",SPHEROID["GRS_1980",6378137.0,298.257222101]],PRIMEM["Greenwich",0.0],'
                'UNIT["Degree",0.0174532925199433]],PROJECTION["Lambert_Conformal_Conic"],'
                'PARAMETER["False_Easting",984250.0],PARAMETER["False_Northing",0.0],PARAMETER["Central_Meridian",-74.0],'
                'PARAMETER["Standard_Parallel_1",40.66666666666666],PARAMETER["Standard_Parallel_2",41.03333333333333],'
                'PARAMETER["Latitude_Of_Origin",40.16666666666666],UNIT["Foot_US",0.3048006096012192],AUTHORITY["EPSG",6539]]'
            ) 

        else:

            logger.error(f"I dont know what to do with srid {srid}")
            return 1        

        ws_sr["F2"] = -1
        # the inverse of the XY resolution
        # 1 / 0.000328083333333333 = 3048.006096012195121164435877261997627062
        # reminder: This is the size of the squares of the graph paper
        # This is not the real world decimal places being collected and maintained 
        # These extra digits do not significantly change the distance between lines on the paper
        # CCHEN wisely:
        # "I want .5 pounds of Virginia Ham vs I want .5123456789 pounds of Virginia ham"
        # "--> same price" 
        ws_sr["K2"] = "3048.006096012195121164435877261997627062" 
        # previously 
        # ws_sr["K2"] = "3048.006096"  
        ws_sr["N2"] = "10000."
        ws_sr["Q2"] = "10000."

        # Fix spatial reference in DatasetContainers sheet (col E)
        ws_dc = wb["DatasetContainers"]
        for row in ws_dc.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                if cell.value and cell.value != -1:
                    cell.value = 1

        # Fix spatial reference in ObjectClasses sheet (col J)
        ws_oc = wb["ObjectClasses"]
        for row in ws_oc.iter_rows(min_row=2, min_col=10, max_col=10):
            for cell in row:
                if cell.value and cell.value != -1:
                    cell.value = 1

        # Save updated Excel
        logger.info(f"Saving updated Excel to {input_spec_xlsx}")
        if os.path.exists(input_spec_xlsx):
            os.remove(input_spec_xlsx)
        wb.save(input_spec_xlsx)

    if not arcpy.Exists(gdbin.gdb):
        raise RuntimeError(f"Input Geodatabase {gdbin.gdb} not found. Cannot continue.")

    if not arcpy.Exists(input_spec_xlsx):
        raise RuntimeError(f"INPUT SPEC File {input_spec_xlsx} not found. Cannot continue.")
             

    if arcpy.Exists(gdbout.gdb): 
        gdbout.clean()
    
    ## Begin reprojection process
    logger.info(f"Creating Target Geodatabase {gdbout.gdb}")

    # this step generates an empty copy of CSCL
    # globalid columns are present and geodatabase-managed
    # change CorrectedProjection.xlsx Fields sheet globalid 
    # FieldType to esriFieldTypeString
    # FieldEditable to TRUE 
    # Creates a dataset with GLOBALID that is not esri managed

    # if gdbout is an enterprise geodatabase schema
    # GenerateGeodatabaseFromExcel throws
    # ERROR 087396: Not a valid SDE workspace. 
    arcpy.topographic.GenerateGeodatabaseFromExcel(input_spec_xlsx
                                                  ,gdbout.gdb)

    logger.info(f"Creating object mapping database for data Load Process - {object_map_gdb.gdb}")
    
    if arcpy.Exists(object_map_gdb.gdb):
        object_map_gdb.clean()

    # no globalids in FieldMapping in the original version
    # adding globalid as type text to gdbout does not create a globalid column in the 
    # cross reference geodatabase (object_map_gdb)
    
    arcpy.topographic.CreateCrossReferenceGeodatabase(gdbin.gdb
                                                     ,gdbout.gdb
                                                     ,object_map_gdb.gdb)

    logger.info(f"Loading Data from {gdbin.gdb} into {gdbout.gdb}")
    

    # https://pro.arcgis.com/en/pro-app/latest/tool-reference/environment-settings/preserve-globalids.htm
    # arcpy.env.preserveGlobalIds = True
    # this does nothing in LoadData
    # and causes table loads to fail due to 
    #   arcgisscripting.ExecuteError: ERROR 003340: The target dataset must have a 
    #   GlobalID field with a unique index in order to use the Preserve GlobalIDs 
    #   geoprocessing environment setting.
    # ie must be enterprise geodatabase as documented 

    arcpy.topographic.LoadData(object_map_gdb.gdb
                              ,gdbin.gdb
                              ,gdbout.gdb)

    # https://pro.arcgis.com/en/pro-app/latest/tool-reference/topographic-production/load-data.htm
    # LoadData row_level_errors = True (default)
    # Errors that occur during individual row-level inserts will be logged.

    # Walk through the objects in the geodatabase
    logger.info("Reading GDB Items to find Tables and Relationship Classes that need to be populated")
    gdb_items = {}
    tables = {}
    counts = {}

    for dirpath, dirnames, filenames in arcpy.da.Walk(gdbin.gdb, topdown=True):
        for filename in filenames:
            # Get the full path of the object
            object_path = os.path.join(dirpath, filename)

            # Get the name of the object
            object_name = os.path.splitext(filename)[0]

            # Get the type of the object
            object_type = arcpy.Describe(object_path).datasetType
            logger.debug(f"FileName:{filename} - object_name:{object_name} - object_type:{object_type}")
            dir_parts = os.path.split(dirpath)
            logger.debug(f"DirParts = {dir_parts}")
            if ".gdb" not in dir_parts[-1]:
                target_gdb = os.path.join(gdbout.gdb, dir_parts[-1])
            else:
                target_gdb = gdbout.gdb

            gdb_items[filename] = (object_type, dirpath, target_gdb)

            if counts.get(object_type, False):
                counts[object_type] += 1
            else:
                counts[object_type] = 1

            #logger.debug((f"FileName:{filename} - object_name:{object_type} - object_type:{gdbout.gdb}"))
            logger.debug((f"Name = {filename} - Type = {object_type} - TargetGDB = {target_gdb}"))
                #f"Name = {pad_field(filename, 40)} - Type = {pad_field(object_type, 30)} - TargetGDB = {target_gdb}"

            if object_type == "Table":
                tables[filename] = (dirpath, target_gdb)

            logger.debug(f"DIRPath: {dirpath}   Object Type: {object_type}    Object Name: {filename}")
    logger.info("-- Item counts")
    total_items = 0
    for tbl_name in counts:
        #logger.info(f"----- {pad_field(tbl_name, 17)} - {counts[tbl_name]}")
        logger.info(f"----- {tbl_name} - {counts[tbl_name]}")
        total_items += counts[tbl_name]
    logger.info(f"-- TOTAL NUMBER OF ITEMS = {total_items}")

    logger.info("Creating Relationship Classes")
    for dirpath, dirnames, filenames in arcpy.da.Walk(gdbin.gdb, datatype=["RelationshipClass"], topdown=True):
        for filename in filenames:
            # Get the full path of the object
            object_path = os.path.join(dirpath, filename)

            # Get the name of the object
            object_name = os.path.splitext(filename)[0]

            # Get the type of the object
            object_type = arcpy.Describe(object_path).datasetType

            dir_parts = os.path.split(dirpath)
            logger.debug(f"DirParts = {dir_parts}")
            if ".gdb" not in dir_parts[-1]:
                target_gdb = os.path.join(gdbout.gdb, dir_parts[-1])
            else:
                target_gdb = gdbout.gdb

            if object_type == "RelationshipClass":
                logger.info(f"{filename} - {object_type} - {dirpath}  ")
                #logger.info(f"{pad_field(filename, 50)} - {pad_field(object_type, 15)} - {dirpath}  ")
                logger.info(f"{filename} - {object_type} - {dirpath}  ")
                rel_class = os.path.join(dirpath, filename)
                desc = arcpy.da.Describe(rel_class)
                logger.info(f"--- out_relationship_class  = {desc['name']} - GDB = {target_gdb}")
                out_relationship_class = os.path.join(target_gdb, desc["name"])

                if len(desc["originClassNames"]) > 1:
                    raise RuntimeError(
                        f"Multiple ORIGIN TABLES encountered - Fix Processing - Assumes 1 - {desc['originClassNames']}"
                    )

                path_list = gdb_items.get(desc["originClassNames"][0], False)
                if not path_list:
                    raise RuntimeError(f"Origin item not found in GDB - {desc['originClassNames']}")
                else:
                    path = path_list[2]

                logger.info(f"--- origin_table            = {desc['originClassNames'][0]} - path = {path}")
                origin_table = os.path.join(path, desc["originClassNames"][0])

                if len(desc["destinationClassNames"]) > 1:
                    raise RuntimeError(
                        f"Multiple DESTINATION TABLES encountered - Fix Processing - Assumes 1 - {desc['destinationClassNames']}"
                    )

                path_list = gdb_items.get(desc["destinationClassNames"][0], False)
                if not path_list:
                    raise RuntimeError(f"Destination item not found in GDB - {desc['destinationClassNames']}")
                else:
                    path = path_list[2]

                logger.info(f"--- destination_table       = {desc['destinationClassNames'][0]} - path = {path}")
                destination_table = os.path.join(path, desc["destinationClassNames"][0])

                if desc["isComposite"]:
                    rel_type = "COMPOSITE"
                else:
                    rel_type = "SIMPLE"
                logger.info(f"----- relationship_type       = {rel_type}")
                relationship_type = rel_type

                logger.info(f"----- forward_label           = {desc['forwardPathLabel']}")
                forward_label = desc["forwardPathLabel"]

                logger.info(f"----- backward_label          = {desc['backwardPathLabel']}")
                backward_label = desc["backwardPathLabel"]

                logger.info(f"----- message_direction       = {desc['notification']}")
                message_direction = desc["notification"]

                if desc["cardinality"].upper() == "ONETOONE":
                    cardinality = "ONE_TO_ONE"
                elif desc["cardinality"].upper() == "ONETOMANY":
                    cardinality = "ONE_TO_MANY"
                elif desc["cardinality"].upper() == "MANYTOMANY":
                    cardinality = "MANY_TO_MANY"
                else:
                    raise RuntimeError(f"Unexpected Cardinality encountered - {desc['cardinality']} - Program stopped")

                if desc["isAttributed"]:
                    # https://github.com/mattyschell/cscl-migrate/issues/39
                    logger.info('adding attributed relclass {0} {1}'.format(str(dirpath), str(target_gdb)))
                    tables[filename] = (dirpath, target_gdb)

                logger.info(f"----- cardinality             = {cardinality}")

                attributed = "ATTRIBUTED" if desc["isAttributed"] else "NONE"
                logger.info(f"----- attributed              = {attributed}")

                origin_primary = ""
                origin_foreign = ""
                for tbl_name in desc["originClassKeys"]:
                    if tbl_name[1] == "OriginPrimary":
                        origin_primary = tbl_name[0]
                    elif tbl_name[1] == "OriginForeign":
                        origin_foreign = tbl_name[0]
                logger.info(f"----- origin_primary_key      = {origin_primary}")
                origin_primary_key = origin_primary

                logger.info(f"----- origin_foreign_key      = {origin_foreign}")
                origin_foreign_key = origin_foreign

                destination_primary = None
                destination_foreign = None
                for tbl_name in desc["destinationClassKeys"]:
                    if tbl_name[1] == "DestinationPrimary":
                        destination_primary = tbl_name[0]
                    elif tbl_name[1] == "DestinationForeign":
                        destination_foreign = tbl_name[0]

                logger.info(f"----- destination_primary_key = {destination_primary}")
                destination_primary_key = destination_primary

                logger.info(f"----- destination_foreign_key = {destination_foreign}")
                destination_foreign_key = destination_foreign

                # TODO: verify that this works
                # delete_item(out_relationship_class)
                arcpy.management.Delete(out_relationship_class)
                
                arcpy.management.CreateRelationshipClass(
                    origin_table=origin_table,
                    destination_table=destination_table,
                    out_relationship_class=out_relationship_class,
                    relationship_type=relationship_type,
                    forward_label=forward_label,
                    backward_label=backward_label,
                    message_direction=message_direction,
                    cardinality=cardinality,
                    attributed=attributed,
                    origin_primary_key=origin_primary_key,
                    origin_foreign_key=origin_foreign_key,
                    destination_primary_key=destination_primary_key,
                    destination_foreign_key=destination_foreign_key,
                )

    logger.info("Loading Tables and attributed relationship class tables")
    for tbl_name in tables:
        source_gdb, target_gdb = tables[tbl_name]
        source_ds = os.path.join(source_gdb, tbl_name)
        target_ds = os.path.join(target_gdb, tbl_name)
        logger.info(f"{tbl_name} - Copying {source_ds} to {target_ds}")

        arcpy.management.TruncateTable(target_ds)
        arcpy.management.Append(source_ds
                               ,target_ds
                               ,schema_type="TEST")
        #arcpy.management.Append(source_ds, target_ds, schema_type="NO_TEST")
        load_count = int(arcpy.management.GetCount(target_ds)[0])  # type: ignore
        if load_count == 0:
            logger.warning(f"--- {tbl_name} has no records")

    logger.info("Counting Source and Target Feature Class records")
    num_mismatch = 0
    for tbl_name in gdb_items:
        obj_type = gdb_items[tbl_name][0]
        src_db = os.path.join(gdb_items[tbl_name][1], tbl_name)
        tgt_db = os.path.join(gdb_items[tbl_name][2], tbl_name)

        if obj_type.upper() == "FEATURECLASS":
            src_count = int(arcpy.management.GetCount(src_db)[0])  # type: ignore
            tgt_count = int(arcpy.management.GetCount(tgt_db)[0])  # type: ignore

            if src_count == 0:
                logger.warning(f"--- {tbl_name} has no records")

            if src_count != tgt_count:
                logger.warning(f"@@@@ Count mismatch for {tbl_name} - Source = {src_count} - Target = {tgt_count}")
                num_mismatch += 1
    if num_mismatch == 0:
        logger.info("All Feature Class record counts match")
    else:
        logger.warning(f"@@@ {num_mismatch} Feature Class record counts don't match")

    logger.info("Counting Source and Target Table records")
    num_mismatch = 0
    for tbl_name in tables:
        src_db = os.path.join(tables[tbl_name][0], tbl_name)
        tgt_db = os.path.join(tables[tbl_name][1], tbl_name)

        src_count = arcpy.management.GetCount(src_db)[0]
        tgt_count = arcpy.management.GetCount(tgt_db)[0]

        if src_count == 0:
            logger.info(f"-- Warning: {tbl_name} has no records")

        if src_count != tgt_count:
            logger.warning(f"@@@ Count mismatch for {tbl_name} - Source = {src_count} - Target = {tgt_count}")
            num_mismatch += 1
    if num_mismatch == 0:
        logger.info("All Table record counts match")
    else:
        logger.warning(f"@@@ {num_mismatch} Table record counts don't match")

    logger.info("CSCL REPROJECTION COMPLETE")
    return 0

if __name__ == '__main__':

    pingdb   = sys.argv[1]
    poutgdb  = sys.argv[2]
    pworkdir = sys.argv[3]

    if len(sys.argv) == 5:
        poutsrid = sys.argv[4]
    else:
        poutsrid = 2263

    # aka "Topographic Production Tools"
    # aka "ArcGIS Production Mapping"
    requiredextension = 'Foundation'

    localgdbin  = localgdb(pingdb)
    localgdbout = localgdb(poutgdb)
    
    timestr = time.strftime("%Y%m%d-%H%M%S")

    # ..\logs\reprojectgdb-ditcspd1-cscl-20250319-110230.log
    targetlog = os.path.join(os.environ['TARGETLOGDIR'] 
                            ,'reprojectgdb-{0}-{1}-{2}.log'.format(localgdbin.basename
                                                                  ,localgdbout.basename
                                                                  ,timestr))

    logging.basicConfig (
        level=logging.INFO,  
        format='%(asctime)s - %(levelname)s - %(message)s',  
        handlers=[
            logging.FileHandler(targetlog),  # log messages 
            logging.StreamHandler()          # cc print 
        ]
    )

    # Create a logger object
    logger = logging.getLogger(__name__)

    if arcpy.CheckExtension(requiredextension) == "Available":
        arcpy.CheckOutExtension(requiredextension)
    else:
        logger.error("Extension {0} is not available".format(requiredextension))
        sys.exit(1)

    retval = reproject(localgdbin
                      ,localgdbout
                      ,logger
                      ,pworkdir
                      ,poutsrid)
    
    arcpy.CheckInExtension(requiredextension)

    logger.info("Completed call to reproject {0} to {1}".format(localgdbin.gdb
                                                               ,localgdbout.gdb))
    
    sys.exit(retval)
