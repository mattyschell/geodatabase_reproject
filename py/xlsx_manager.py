import os
import arcpy
import openpyxl


class ExcelFile(object):

    def __init__(self
                ,excelfile):

        self.xlsx = excelfile
        self.wb   = None
        self._license_checked_out = False
        self._closed = False

    def checkoutlicense(self):
        if not self._license_checked_out:
            if arcpy.CheckExtension("Foundation") == "Available":
                arcpy.CheckOutExtension("Foundation")
                self._license_checked_out = True
            else:
                raise RuntimeError("Foundation license not available")

    def checkinlicense(self):
        if not self._closed:
            if self._license_checked_out:
                arcpy.CheckInExtension("Foundation")
                self._license_checked_out = False
            self._closed = True

    def exists(self):

        if os.path.exists(self.xlsx):
            return True
        else:
            return False

    def delete(self):
        try:
            os.remove(self.xlsx)
        except FileNotFoundError:
            pass

    def _save(self):
        self.wb.save(self.xlsx)

    def _open(self):
        self.wb = openpyxl.load_workbook(self.xlsx)

    def copy(self
            ,output_xlsx):

        self._open()
        self.wb.save(output_xlsx)

    def generate_from_geodatabase(self
                                 ,gdb):

        arcpy.topographic.GenerateExcelFromGeodatabase(gdb
                                                      ,self.xlsx)
        self.wb = openpyxl.load_workbook(self.xlsx)
    
    def generate_to_geodatabase(self
                               ,gdb):

        # this generates an empty geodatabase
        # globalids if present will be fresh and ESRI-managed 
        arcpy.topographic.GenerateGeodatabaseFromExcel(self.xlsx
                                                      ,gdb)

    def _get_srid_dictionary(self
                            ,srid):

        if srid == 2263:
            srid_dict = {
                'A2': 1,
                'B2': 'NAD_1983_StatePlane_New_York_Long_Island_FIPS_3104_Feet',
                'C2': 2263,
                "D2": (
                    'PROJCS["NAD_1983_StatePlane_New_York_Long_Island_FIPS_3104_Feet",'
                    'GEOGCS["GCS_North_American_1983",'
                    'DATUM["D_North_American_1983",SPHEROID["GRS_1980",6378137.0,298.257222101]],'
                    'PRIMEM["Greenwich",0.0],'
                    'UNIT["Degree",0.0174532925199433]],'
                    'PROJECTION["Lambert_Conformal_Conic"],'
                    'PARAMETER["False_Easting",984250.0],'
                    'PARAMETER["False_Northing",0.0],'
                    'PARAMETER["Central_Meridian",-74.0],'
                    'PARAMETER["Standard_Parallel_1",40.66666666666666],'
                    'PARAMETER["Standard_Parallel_2",41.03333333333333],'
                    'PARAMETER["Latitude_Of_Origin",40.16666666666666],'
                    'UNIT["Foot_US",0.3048006096012192]]'
                ),
                'F2': -1,
                # the inverse of the XY resolution
                # 1 / 0.000328083333333333 = 3048.006096012195121164435877261997627062
                # reminder: This is the size of the squares of the graph paper
                # This is not the real world decimal places being collected and maintained 
                # These extra digits do not meaningfully change the graph paper 
                'K2': '3048.006096012195121164435877261997627062',
                'N2': "10000.",
                'Q2': "10000.",                
            }
        elif srid == 6539: 
            srid_dict = {
                'A2': 1,
                'B2': 'NAD_1983_2011_StatePlane_New_York_Long_Isl_FIPS_3104_Ft_US',
                'C2': 6539,
                "D2": (
                    'PROJCS["NAD_1983_2011_StatePlane_New_York_Long_Isl_FIPS_3104_Ft_US",'
                    'GEOGCS["GCS_NAD_1983_2011",'
                    'DATUM["D_NAD_1983_2011",SPHEROID["GRS_1980",6378137.0,298.257222101]],'
                    'PRIMEM["Greenwich",0.0],'
                    'UNIT["Degree",0.0174532925199433]],'
                    'PROJECTION["Lambert_Conformal_Conic"],'
                    'PARAMETER["False_Easting",984250.0],'
                    'PARAMETER["False_Northing",0.0],'
                    'PARAMETER["Central_Meridian",-74.0],'
                    'PARAMETER["Standard_Parallel_1",40.66666666666666],'
                    'PARAMETER["Standard_Parallel_2",41.03333333333333],'
                    'PARAMETER["Latitude_Of_Origin",40.16666666666666],'
                    'UNIT["Foot_US",0.3048006096012192],'
                    'AUTHORITY["EPSG",6539]]'
                ),
                'F2': -1,
                'K2': '3048.006096012195121164435877261997627062',
                'N2': "10000.",
                'Q2': "10000.",
            }
        else:
            raise ValueError('We have not added SRID {0}'.format(srid))

        return srid_dict  

    def _update_dataset_containers(self):
        
        # update spatial reference in DatasetContainers sheet (col E)
        # all will point at srid 1
        ws_dc = self.wb["DatasetContainers"]
        for row in ws_dc.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                if cell.value and cell.value != -1:
                    cell.value = 1
    
    def _update_object_classes(self):

        # set all object classes to srid 1
        ws_oc = self.wb["ObjectClasses"]
        for row in ws_oc.iter_rows(min_row=2, min_col=10, max_col=10):
            for cell in row:
                if cell.value and cell.value != -1:
                    cell.value = 1

    def _update_spatial_reference(self
                                 ,srid):

        #update spatial reference by overwriting SpatialReferences worksheet
        ws_sr = self.wb["SpatialReferences"]
        # iter_rows stops at no data
        for row in ws_sr.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        # this is the good spatial reference
        for cell, value in self._get_srid_dictionary(srid): 
            ws_sr[cell] = value

    def update_all_spatial_reference(self
                                    ,srid):

        # update_all_spatial_reference will associates all data
        # with one blessed srid
        srid = int(srid)
        self._open()        

        self._update_spatial_reference(srid)
        self._update_dataset_containers()
        self._update_object_classes()
        
        self._save()
    


        